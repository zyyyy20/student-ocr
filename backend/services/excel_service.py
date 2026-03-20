from __future__ import annotations

import re
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelService:
    """
    使用 openpyxl 生成导出 Excel。
    """

    def __init__(self, *, export_dir: Path) -> None:
        self.export_dir = export_dir
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_transcript(self, payload: Dict[str, Any]) -> str:
        """
        输入：前端修改后的成绩 JSON（{"headers": [...], "rows": [...]}）
        输出：生成 Excel 文件名（用于拼接静态下载链接）
        """
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="导出数据格式错误：payload 必须为 JSON 对象")

        title = payload.get("title") or ""
        headers = payload.get("headers") or []
        rows = payload.get("rows") or []
        
        if not isinstance(headers, list) or not isinstance(rows, list):
            raise HTTPException(status_code=400, detail="导出数据格式错误：headers/rows 结构不正确")

        wb = Workbook()
        ws = wb.active
        ws.title = "学生成绩单"

        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin_side = Side(style="thin", color="B8C4CE")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        row_offset = 0

        # 0. 写入标题（可选）
        if isinstance(title, str) and title.strip():
            row_offset = 1
            title_value = title.strip()
            ws.cell(row=1, column=1, value=title_value).font = title_font
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
            if headers:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))

        # 1. 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1 + row_offset, column=col_idx, value=str(header))
            cell.font = bold
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = cell_border

        # 2. 写入数据行
        for row_idx, row_item in enumerate(rows, start=2 + row_offset):
            # row_item 可能是 {"values": {...}, "confidences": {...}}
            values = row_item.get("values", {})
            if not isinstance(values, dict):
                continue
                
            for col_idx, header in enumerate(headers, start=1):
                val = values.get(header, "")
                # 尝试转数字以便 Excel 统计
                num_val = self._parse_score(val)
                final_val = num_val if num_val is not None else val

                cell = ws.cell(row=row_idx, column=col_idx, value=final_val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = cell_border

        header_row = 1 + row_offset
        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = ws.dimensions
        self._autosize_columns(ws, headers, rows)

        filename = f"class_transcript_{uuid.uuid4().hex}.xlsx"
        out_path = self.export_dir / filename
        wb.save(str(out_path))
        return filename

    def _autosize_columns(self, ws, headers: List[Any], rows: List[Dict[str, Any]]) -> None:
        for col_idx, header in enumerate(headers, start=1):
            values = [str(header)]
            for row in rows:
                row_values = row.get("values", {})
                if isinstance(row_values, dict):
                    values.append(str(row_values.get(header, "")))
            max_len = max((len(v) for v in values), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 30)

    @staticmethod
    def _parse_score(v: Any) -> Optional[float | int]:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return int(v) if isinstance(v, int) or abs(v - int(v)) < 1e-9 else float(v)

        s = str(v).strip()
        if not re.match(r"^-?\d+(\.\d+)?$", s):
            return None
            
        try:
            num = float(s)
            return int(num) if abs(num - int(num)) < 1e-9 else round(num, 2)
        except Exception:
            return None
