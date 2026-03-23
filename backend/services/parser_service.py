from __future__ import annotations

import io
import re
import xml.etree.ElementTree as ET
import base64
from typing import Any, Dict, List, Optional, Sequence, Tuple

import cv2
import numpy as np
from fastapi import HTTPException
from openpyxl import load_workbook

from backend.services.ocr_service import OCRItem, OCRService


class FileParserService:
    """
    文件解析统一入口：
    - 支持班级成绩表解析（动态表头，多行数据）
    - PNG/JPG：PP-Structure 表格识别 + 全图 OCR 兜底
    - SVG：优先解析 <text> 文本节点；失败则转图片 OCR
    - XLSX：直接读取表格数据
    输出结构：{"headers": [...], "rows": [{"values": {...}, "confidences": {...}}, ...]}
    """

    def __init__(self, *, ocr_service: OCRService) -> None:
        self.ocr_service = ocr_service

    def parse(self, *, filename: str, content_type: str, data: bytes) -> Dict[str, Any]:
        ext = self._guess_ext(filename, content_type, data)
        if ext in {".png", ".jpg", ".jpeg"}:
            return self._parse_image(data)
        if ext == ".svg":
            return self._parse_svg(data)
        if ext == ".xlsx":
            return self._parse_xlsx(data)
        raise HTTPException(status_code=400, detail=f"不支持的文件类型：{ext}（仅支持 PNG/JPG/SVG/XLSX）")

    @staticmethod
    def _guess_ext(filename: str, content_type: str, data: bytes) -> str:
        name = (filename or "").lower()
        for ext in (".png", ".jpg", ".jpeg", ".svg", ".xlsx"):
            if name.endswith(ext):
                return ext
        ct = (content_type or "").lower()
        if "svg" in ct:
            return ".svg"
        if "excel" in ct or "spreadsheetml" in ct:
            return ".xlsx"
        if "png" in ct:
            return ".png"
        if "jpeg" in ct or "jpg" in ct:
            return ".jpg"
        if data.startswith(b"\x89PNG\r\n\x1a\n"):
            return ".png"
        if data[:3] == b"\xff\xd8\xff":
            return ".jpg"
        if data.startswith(b"PK\x03\x04"):
            return ".xlsx"
        head = data[:500].lstrip().lower()
        if head.startswith(b"<svg") or b"<svg" in head:
            return ".svg"
        return ""

    def _parse_image(self, data: bytes) -> Dict[str, Any]:
        image_bgr = self.ocr_service.decode_image(data)
        preprocess_result = self.ocr_service.preprocess_with_context(image_bgr)
        image_bgr = preprocess_result["image"]
        transform_context = preprocess_result["context"]
        image_size = transform_context["original_size"] if transform_context else None

        # 先做一次文本 OCR（后续置信度匹配、表格重建都复用）
        ocr_items = self.ocr_service.recognize_text(image_bgr, preprocess=False)
        meta = self._extract_transcript_meta(ocr_items)
        image_bgr, transform_context, focused = self._focus_table_region(
            image_bgr,
            ocr_items,
            transform_context,
        )
        if focused:
            ocr_items = self.ocr_service.recognize_text(image_bgr, preprocess=False)
            meta = self._extract_transcript_meta(ocr_items)

        table = None
        try:
            table = self.ocr_service.recognize_table(image_bgr, ocr_items=ocr_items, preprocess=False)
        except Exception:
            table = None

        if table and len(table) >= 2:
            result = self._extract_class_grid(
                table,
                image_bgr=image_bgr,
                ocr_items=ocr_items,
                default_conf=0.85,
                image_size=image_size,
                transform_context=transform_context,
            )
            result["processed_preview"] = self._encode_preview_image(image_bgr)
            if meta:
                result["meta"] = meta
                if not result.get("title") and meta.get("标题"):
                    result["title"] = str(meta["标题"]).strip()
            return result

        # 兜底：无法识别表格结构时，返回单列文本列表
        result = self._extract_from_text_lines_fallback(
            [it.text for it in ocr_items],
            meta=meta,
            image_size=image_size,
            ocr_items=ocr_items,
        )
        result["processed_preview"] = self._encode_preview_image(image_bgr)
        return result

    @staticmethod
    def _encode_preview_image(image_bgr: np.ndarray) -> str:
        ok, buf = cv2.imencode(".png", image_bgr)
        if not ok:
            return ""
        return f"data:image/png;base64,{base64.b64encode(buf.tobytes()).decode('ascii')}"

    def _focus_table_region(
        self,
        image_bgr: np.ndarray,
        ocr_items: Sequence[OCRItem],
        transform_context: Optional[Dict[str, Any]],
    ) -> Tuple[np.ndarray, Optional[Dict[str, Any]], bool]:
        bbox = self._estimate_focus_bbox(image_bgr, ocr_items)
        if bbox is None:
            return image_bgr, transform_context, False

        x1, y1, x2, y2 = bbox
        cropped = image_bgr[y1 : y2 + 1, x1 : x2 + 1]
        if cropped.size == 0:
            return image_bgr, transform_context, False

        if transform_context is None:
            return cropped, transform_context, True

        crop_matrix = np.array([[1, 0, -x1], [0, 1, -y1], [0, 0, 1]], dtype=np.float32)
        forward_matrix = crop_matrix @ np.array(transform_context["forward_matrix"], dtype=np.float32)
        new_context = {
            "forward_matrix": forward_matrix,
            "inverse_matrix": np.linalg.inv(forward_matrix),
            "original_size": transform_context["original_size"],
            "processed_size": {"width": int(cropped.shape[1]), "height": int(cropped.shape[0])},
        }
        return cropped, new_context, True

    def _estimate_focus_bbox(
        self,
        image_bgr: np.ndarray,
        ocr_items: Sequence[OCRItem],
    ) -> Optional[Tuple[int, int, int, int]]:
        if image_bgr is None or not ocr_items:
            return None

        h, w = image_bgr.shape[:2]
        footer_keywords = ("班主任", "老师", "日期", "年月日")
        candidates: List[Tuple[float, float, float, float]] = []
        footer_top: Optional[float] = None

        for item in ocr_items:
            box = self.ocr_service._xyxy(item.box)
            if not box:
                continue
            x1, y1, x2, y2 = box
            text = str(item.text or "").strip()
            if any(keyword in text for keyword in footer_keywords):
                footer_top = y1 if footer_top is None else min(footer_top, y1)
                continue
            if (x2 - x1) * (y2 - y1) < 60:
                continue
            candidates.append((x1, y1, x2, y2))

        if len(candidates) < 6:
            return None

        xs1 = [b[0] for b in candidates]
        ys1 = [b[1] for b in candidates]
        xs2 = [b[2] for b in candidates]
        ys2 = [b[3] for b in candidates]
        x1 = max(0, int(np.percentile(xs1, 3)) - 24)
        y1 = max(0, int(np.percentile(ys1, 2)) - 24)
        x2 = min(w - 1, int(max(xs2)) + 44)
        y2 = min(h - 1, int(np.percentile(ys2, 97)) + 40)

        if footer_top is not None and footer_top > h * 0.6:
            y2 = min(y2, int(footer_top) - 12)

        if (x2 - x1) < w * 0.35 or (y2 - y1) < h * 0.2:
            return None

        return x1, y1, x2, y2

    def _parse_svg(self, data: bytes) -> Dict[str, Any]:
        # 1) 优先解析 SVG 文本节点
        try:
            texts = self._extract_svg_text_nodes(data)
        except Exception:
            texts = []

        if texts:
             # SVG 文本通常是散乱的，很难直接构建二维表格
             # 这里简单处理：转为单列数据
             return self._extract_from_text_lines_fallback(texts)

        # 2) 兜底：SVG 转图片再 OCR
        try:
            import cairosvg
            png_bytes = cairosvg.svg2png(bytestring=data)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"SVG 解析失败且无法转图片：{e}") from e

        return self._parse_image(png_bytes)

    def _parse_xlsx(self, data: bytes) -> Dict[str, Any]:
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"XLSX 读取失败：{e}") from e

        ws = wb.active
        rows = []
        for r in ws.iter_rows(values_only=True):
            row = [("" if v is None else str(v).strip()) for v in r]
            # 只要这一行有内容，就保留（哪怕是空行中间的）
            if any(c for c in row):
                rows.append(row)
        
        result = self._extract_class_grid(
            rows,
            image_bgr=None,
            ocr_items=[],
            default_conf=1.0,
            image_size=None,
        )
        result.setdefault("meta", {})
        return result

    @staticmethod
    def _extract_svg_text_nodes(data: bytes) -> List[str]:
        try:
            root = ET.fromstring(data.decode("utf-8", errors="ignore"))
        except Exception:
            root = ET.fromstring(data)

        out: List[str] = []
        for el in root.iter():
            tag = el.tag.split("}")[-1].lower()
            if tag == "text":
                txt = " ".join(t.strip() for t in el.itertext() if t and t.strip()).strip()
                if txt:
                    out.append(txt)
        return out

    def _extract_transcript_meta(self, ocr_items: Sequence[OCRItem]) -> Dict[str, Any]:
        meta: Dict[str, Any] = {}
        lines = [str(item.text).strip() for item in ocr_items if str(item.text).strip()]
        if not lines:
            return meta

        title = next((line for line in lines[:5] if any(k in line for k in ("成绩单", "成绩表"))), "")
        if title:
            meta["标题"] = title

        field_patterns = {
            "姓名": r"姓名[:：]?\s*([^\s:：]+)",
            "学号": r"学号[:：]?\s*([A-Za-z0-9]+)",
            "班级": r"班级[:：]?\s*([^\s]+)",
            "学院": r"学院[:：]?\s*([^\s]+)",
            "学期": r"(?:学期|学年学期)[:：]?\s*([^\s]+)",
        }

        for line in lines:
            compact = re.sub(r"\s+", "", line)
            for field, pattern in field_patterns.items():
                if field in meta:
                    continue
                m = re.search(pattern, compact)
                if m:
                    meta[field] = m.group(1)

        return meta

    def _extract_class_grid(
        self,
        grid: List[List[str]],
        *,
        image_bgr: Optional[np.ndarray] = None,
        ocr_items: Sequence[OCRItem],
        default_conf: float,
        image_size: Optional[Dict[str, int]],
        transform_context: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        通用表格解析逻辑：
        - 第一行非空行作为表头（Headers）
        - 后续行作为数据（Rows）
        - 自动转换数字
        - 匹配置信度
        """
        if not grid:
            return {"headers": [], "rows": []}

        # 1. 寻找表头行（假设第一行就是，或者根据关键词探测）
        header_row_idx = 0
        title_text: Optional[str] = None
        for i, row in enumerate(grid[:5]):
            joined = "".join(row)
            non_empty = sum(1 for c in row if (c or "").strip())
            # 标题行通常只有 1 个非空单元格，且包含“成绩表/成绩单”
            if non_empty <= 2 and any(k in joined for k in ("成绩表", "成绩单")):
                title_text = joined.strip()
                continue
            # 如果包含常见表头关键词，可能是表头
            if any(k in joined for k in ("姓名", "班级", "学号", "科目", "成绩", "分数", "等级")):
                header_row_idx = i
                break

        # 若第一行是标题（单格居中），而第二行更像表头，则跳过标题
        if len(grid) >= 2:
            first = grid[header_row_idx]
            second = grid[header_row_idx + 1] if header_row_idx + 1 < len(grid) else []
            first_non_empty = sum(1 for c in first if (c or "").strip())
            second_non_empty = sum(1 for c in second if (c or "").strip())
            first_joined = "".join(first)
            second_joined = "".join(second)
            if (
                first_non_empty <= 2
                and (any(k in first_joined for k in ("成绩表", "成绩单")) or first_non_empty == 1)
                and second_non_empty >= 2
                and any(k in second_joined for k in ("姓名", "班级", "学号", "科目", "成绩", "分数", "等级"))
            ):
                title_text = first_joined.strip() if first_joined.strip() else title_text
                header_row_idx = header_row_idx + 1
        
        headers_raw = self._merge_multirow_headers(grid, header_row_idx)
        headers = [self._normalize_header_name(h.strip()) for h in headers_raw]
        header_boxes = self._extract_header_boxes(
            headers_raw, headers, ocr_items, image_size, transform_context
        )

        # 表头可能存在漏检/空白：为保证列对齐，补全为空表头列命名
        for i in range(len(headers)):
            if not headers[i]:
                headers[i] = f"列{i+1}"

        # 避免表头重复导致覆盖：追加序号确保唯一
        seen: Dict[str, int] = {}
        for i, h in enumerate(headers):
            if h not in seen:
                seen[h] = 1
                continue
            seen[h] += 1
            headers[i] = f"{h}_{seen[h]}"
        
        # 2. 提取数据行
        rows_out = []
        start_row = header_row_idx + 1
        
        for r_idx in range(start_row, len(grid)):
            row_raw = grid[r_idx]
            # 补齐或截断
            if len(row_raw) < len(headers):
                row_raw += [""] * (len(headers) - len(row_raw))
            else:
                row_raw = row_raw[:len(headers)]
            
            # 如果全空则跳过
            if not any(c.strip() for c in row_raw):
                continue
                
            values = {}
            confidences = {}
            boxes = {}
            
            for c_idx, cell_val in enumerate(row_raw):
                key = headers[c_idx]
                
                # 尝试转数字
                val_clean = cell_val.strip()
                val_num = self._parse_score(val_clean)
                final_val = val_num if val_num is not None else val_clean
                
                values[key] = final_val
                
                # 计算置信度
                conf = default_conf
                if ocr_items:
                    # 尝试匹配文本内容获取置信度
                    matched = OCRService.best_match_confidence(val_clean, ocr_items)
                    if matched is not None:
                        conf = matched
                    matched_box = OCRService.best_match_box(val_clean, ocr_items)
                    normalized_box = self._normalize_polygon(
                        matched_box, image_size, transform_context
                    )
                    if normalized_box is not None:
                        boxes[key] = normalized_box
                
                confidences[key] = round(float(conf), 4)
            
            row_payload = {"values": values, "confidences": confidences}
            if boxes:
                row_payload["boxes"] = boxes
            rows_out.append(row_payload)
            
        # 3. 剔除“全空的补位列”（常见于合并单元格导致的空列）
        if rows_out:
            non_empty_counts: Dict[str, int] = {h: 0 for h in headers}
            for row in rows_out:
                vals = row.get("values", {})
                for h in headers:
                    v = vals.get(h, "")
                    if str(v).strip() != "":
                        non_empty_counts[h] += 1

            threshold = max(1, int(len(rows_out) * 0.05))
            drop_headers = {
                h for h in headers
                if re.match(r"^列\d+$", h) and non_empty_counts.get(h, 0) <= threshold
            }
            if drop_headers:
                headers = [h for h in headers if h not in drop_headers]
                for row in rows_out:
                    row["values"] = {k: v for k, v in row["values"].items() if k in headers}
                    row["confidences"] = {k: v for k, v in row["confidences"].items() if k in headers}
                    if "boxes" in row:
                        row["boxes"] = {k: v for k, v in row["boxes"].items() if k in headers}

        rows_out = self._merge_fragmented_rows(headers, rows_out)
        header_boxes, rows_out = self._estimate_cell_regions(headers, header_boxes, rows_out, transform_context)
        rows_out = self._recover_numeric_cells(
            headers,
            rows_out,
            ocr_items,
            image_bgr,
            transform_context,
        )
        rows_out = self._normalize_transcript_rows(headers, rows_out)

        result: Dict[str, Any] = {
            "headers": headers,
            "rows": rows_out,
        }
        if header_boxes:
            result["header_boxes"] = header_boxes
        if title_text:
            result["title"] = title_text
        return result

    @staticmethod
    def _normalize_header_name(name: str) -> str:
        normalized = re.sub(r"\s+", "", name or "")
        alias_map = {
            "课程名称": "课程",
            "科目": "课程",
            "课程名": "课程",
            "分数": "成绩",
            "总评成绩": "成绩",
            "考试成绩": "成绩",
            "课程成绩": "成绩",
            "学年学期": "学期",
            "班别": "班级",
        }
        return alias_map.get(normalized, normalized)

    def _merge_multirow_headers(self, grid: List[List[str]], header_row_idx: int) -> List[str]:
        headers = list(grid[header_row_idx])
        if header_row_idx == 0:
            return headers

        prev_row = grid[header_row_idx - 1]
        if sum(1 for c in prev_row if (c or "").strip()) <= 1:
            return headers

        merged: List[str] = []
        max_len = max(len(prev_row), len(headers))
        for idx in range(max_len):
            top = prev_row[idx].strip() if idx < len(prev_row) else ""
            bottom = headers[idx].strip() if idx < len(headers) else ""
            if top and bottom and top != bottom:
                merged.append(f"{top}_{bottom}")
            else:
                merged.append(bottom or top)
        return merged

    def _merge_fragmented_rows(
        self,
        headers: Sequence[str],
        rows_out: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        if len(rows_out) < 2:
            return rows_out

        merged_rows: List[Dict[str, Any]] = []
        idx = 0
        while idx < len(rows_out):
            current = rows_out[idx]
            if idx + 1 >= len(rows_out):
                merged_rows.append(current)
                break

            nxt = rows_out[idx + 1]
            current_values = current.get("values", {})
            next_values = nxt.get("values", {})
            current_non_empty = [h for h in headers if str(current_values.get(h, "")).strip()]
            next_non_empty = [h for h in headers if str(next_values.get(h, "")).strip()]

            if self._should_merge_fragmented_rows(headers, current_non_empty, next_non_empty):
                merged = {
                    "values": dict(current_values),
                    "confidences": dict(current.get("confidences", {})),
                    "boxes": dict(current.get("boxes", {})),
                }
                for header in headers:
                    if str(merged["values"].get(header, "")).strip():
                        continue
                    next_value = next_values.get(header, "")
                    if str(next_value).strip():
                        merged["values"][header] = next_value
                        conf = nxt.get("confidences", {}).get(header)
                        if conf is not None:
                            merged["confidences"][header] = conf
                        box = nxt.get("boxes", {}).get(header)
                        if box is not None:
                            merged["boxes"][header] = box
                if not merged["boxes"]:
                    merged.pop("boxes")
                merged_rows.append(merged)
                idx += 2
                continue

            merged_rows.append(current)
            idx += 1

        return merged_rows

    @staticmethod
    def _should_merge_fragmented_rows(
        headers: Sequence[str],
        current_non_empty: Sequence[str],
        next_non_empty: Sequence[str],
    ) -> bool:
        if not current_non_empty or not next_non_empty:
            return False
        overlap = set(current_non_empty) & set(next_non_empty)
        if overlap:
            return False
        if len(current_non_empty) > max(3, len(headers) // 2):
            return False
        if len(next_non_empty) > max(3, len(headers) // 2):
            return False

        current_idx = [headers.index(h) for h in current_non_empty if h in headers]
        next_idx = [headers.index(h) for h in next_non_empty if h in headers]
        if not current_idx or not next_idx:
            return False

        current_max = max(current_idx)
        next_min = min(next_idx)

        # 两行字段互补且后一行落在更靠后的列，判定为被拆开的同一条记录
        return next_min >= current_max and (len(current_non_empty) + len(next_non_empty)) <= len(headers)

    def _normalize_transcript_rows(
        self, headers: Sequence[str], rows_out: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        normalized_rows: List[Dict[str, Any]] = []
        for row in rows_out:
            values = dict(row.get("values", {}))
            confidences = dict(row.get("confidences", {}))
            boxes = dict(row.get("boxes", {}))

            if "学号" in values:
                values["学号"] = re.sub(r"\D+", "", str(values["学号"]))
            if "学分" in values:
                credit = self._parse_score(values["学分"])
                values["学分"] = credit if credit is not None else str(values["学分"]).strip()
            if "成绩" in values:
                score = self._parse_score(values["成绩"])
                values["成绩"] = score if score is not None else str(values["成绩"]).strip()
            if "课程" in values:
                values["课程"] = str(values["课程"]).strip()

            # 排除明显是元信息残留的行
            joined = "".join(str(values.get(h, "")).strip() for h in headers)
            if any(k in joined for k in ("姓名", "学号", "班级", "学院")) and len(headers) <= 3:
                continue
            if any(k in joined for k in ("班主任", "老师", "2024年", "2025年", "2026年")):
                continue

            row_payload = {"values": values, "confidences": confidences}
            if boxes:
                row_payload["boxes"] = boxes
            normalized_rows.append(row_payload)

        return normalized_rows

    def _extract_header_boxes(
        self,
        headers_raw: Sequence[str],
        headers: Sequence[str],
        ocr_items: Sequence[OCRItem],
        image_size: Optional[Dict[str, int]],
        transform_context: Optional[Dict[str, Any]],
    ) -> Dict[str, Dict[str, float]]:
        boxes: Dict[str, Dict[str, float]] = {}
        for raw, normalized in zip(headers_raw, headers):
            query = str(raw or "").strip() or str(normalized or "").strip()
            if not query:
                continue
            box = OCRService.best_match_box(query, ocr_items)
            normalized_box = self._normalize_polygon(box, image_size, transform_context)
            if normalized_box is not None:
                boxes[normalized] = normalized_box
        return boxes

    def _estimate_cell_regions(
        self,
        headers: Sequence[str],
        header_boxes: Dict[str, Dict[str, Any]],
        rows_out: List[Dict[str, Any]],
        transform_context: Optional[Dict[str, Any]],
    ) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
        coord_key = "processed_points" if transform_context else "points"
        samples: List[Tuple[str, int, Dict[str, Any]]] = []
        for col_idx, header in enumerate(headers):
            box = header_boxes.get(header)
            if box and box.get(coord_key):
                samples.append((header, -1, box))
        for row_idx, row in enumerate(rows_out):
            for col_idx, header in enumerate(headers):
                box = row.get("boxes", {}).get(header)
                if box and box.get(coord_key):
                    samples.append((header, row_idx, box))

        if len(samples) < 4:
            return header_boxes, rows_out

        origin, x_axis, y_axis = self._estimate_table_axes([box for _, _, box in samples], coord_key)
        if origin is None or x_axis is None or y_axis is None:
            return header_boxes, rows_out

        col_centers: Dict[str, List[float]] = {header: [] for header in headers}
        row_centers: Dict[int, List[float]] = {-1: []}
        for idx in range(len(rows_out)):
            row_centers[idx] = []

        for header, row_idx, box in samples:
            center = self._polygon_center(box, coord_key)
            col_centers[header].append(float(np.dot(center, x_axis)))
            row_centers[row_idx].append(float(np.dot(center, y_axis)))

        col_positions = [self._safe_median(col_centers[h]) for h in headers]
        row_keys = [-1] + list(range(len(rows_out)))
        row_positions = [self._safe_median(row_centers[k]) for k in row_keys]
        if any(v is None for v in col_positions) or any(v is None for v in row_positions):
            return header_boxes, rows_out

        col_bounds = self._positions_to_bounds([float(v) for v in col_positions])
        row_bounds = self._positions_to_bounds([float(v) for v in row_positions])
        col_bounds = self._refine_col_bounds(headers, col_bounds, header_boxes, rows_out, coord_key)
        row_bounds = self._refine_row_bounds(row_bounds, header_boxes, rows_out, coord_key)

        new_header_boxes: Dict[str, Dict[str, Any]] = {}
        for col_idx, header in enumerate(headers):
            new_header_boxes[header] = self._bounds_to_polygon(
                origin,
                x_axis,
                y_axis,
                col_bounds[col_idx],
                col_bounds[col_idx + 1],
                row_bounds[0],
                row_bounds[1],
                transform_context,
                coord_key,
            )

        new_rows: List[Dict[str, Any]] = []
        for row_idx, row in enumerate(rows_out):
            row_copy = dict(row)
            new_boxes: Dict[str, Dict[str, Any]] = {}
            for col_idx, header in enumerate(headers):
                new_boxes[header] = self._bounds_to_polygon(
                    origin,
                    x_axis,
                    y_axis,
                    col_bounds[col_idx],
                    col_bounds[col_idx + 1],
                    row_bounds[row_idx + 1],
                    row_bounds[row_idx + 2],
                    transform_context,
                    coord_key,
                )
            if new_boxes:
                row_copy["boxes"] = new_boxes
            new_rows.append(row_copy)

        return new_header_boxes, new_rows

    def _recover_numeric_cells(
        self,
        headers: Sequence[str],
        rows_out: List[Dict[str, Any]],
        ocr_items: Sequence[OCRItem],
        image_bgr: Optional[np.ndarray],
        transform_context: Optional[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        if not rows_out or not ocr_items:
            return rows_out

        coord_key = "processed_points" if transform_context else "points"
        column_examples: Dict[str, List[float]] = {header: [] for header in headers}
        for row in rows_out:
            values = row.get("values", {})
            for header in headers:
                value = values.get(header, "")
                parsed = self._parse_score(value)
                if parsed is not None:
                    column_examples[header].append(float(parsed))

        recovered_rows: List[Dict[str, Any]] = []
        for row in rows_out:
            row_copy = {
                "values": dict(row.get("values", {})),
                "confidences": dict(row.get("confidences", {})),
                "boxes": dict(row.get("boxes", {})),
            }
            for header in headers:
                if not self._is_numeric_column(header):
                    continue
                current_value = str(row_copy["values"].get(header, "")).strip()
                if current_value:
                    continue
                cell_box = row_copy["boxes"].get(header)
                if not cell_box:
                    continue
                candidate = self._extract_text_from_cell(
                    cell_box,
                    coord_key,
                    ocr_items,
                    transform_context,
                )
                if not candidate and image_bgr is not None:
                    candidate = self._run_local_cell_ocr(
                        image_bgr,
                        cell_box,
                        coord_key,
                        transform_context,
                    )
                if not candidate:
                    continue
                repaired = self._repair_numeric_candidate(candidate, column_examples.get(header, []))
                if repaired is None:
                    continue
                row_copy["values"][header] = repaired
                row_copy["confidences"][header] = round(
                    OCRService.best_match_confidence(str(candidate), ocr_items) or 0.72,
                    4,
                )
            if not row_copy["boxes"]:
                row_copy.pop("boxes")
            recovered_rows.append(row_copy)
        return recovered_rows

    def _extract_text_from_cell(
        self,
        cell_box: Dict[str, Any],
        coord_key: str,
        ocr_items: Sequence[OCRItem],
        transform_context: Optional[Dict[str, Any]],
    ) -> str:
        pts = self._points_array(cell_box, coord_key)
        if pts is None:
            return ""
        x1 = float(np.min(pts[:, 0]))
        y1 = float(np.min(pts[:, 1]))
        x2 = float(np.max(pts[:, 0]))
        y2 = float(np.max(pts[:, 1]))
        pad_x = max((x2 - x1) * 0.08, 0.005)
        pad_y = max((y2 - y1) * 0.12, 0.005)
        tokens: List[Tuple[float, str]] = []
        for item in ocr_items:
            bb = self.ocr_service._xyxy(item.box)
            if not bb:
                continue
            cx = (bb[0] + bb[2]) / 2.0
            cy = (bb[1] + bb[3]) / 2.0
            if coord_key == "processed_points" and transform_context is not None:
                processed_size = transform_context["processed_size"]
                width = max(1.0, float(processed_size["width"]))
                height = max(1.0, float(processed_size["height"]))
                cx_norm = cx / width
                cy_norm = cy / height
            else:
                original_size = (transform_context or {}).get("original_size", {})
                width = max(1.0, float(original_size.get("width", 1.0)))
                height = max(1.0, float(original_size.get("height", 1.0)))
                cx_norm = cx / width
                cy_norm = cy / height
            text = str(item.text or "").strip()
            if not text:
                continue
            if x1 - pad_x <= cx_norm <= x2 + pad_x and y1 - pad_y <= cy_norm <= y2 + pad_y:
                tokens.append((cx_norm, text))

        if not tokens:
            return ""
        tokens.sort(key=lambda item: item[0])
        return "".join(text for _, text in tokens).strip()

    def _run_local_cell_ocr(
        self,
        image_bgr: np.ndarray,
        cell_box: Dict[str, Any],
        coord_key: str,
        transform_context: Optional[Dict[str, Any]],
    ) -> str:
        if image_bgr is None:
            return ""
        pts = self._points_array(cell_box, coord_key)
        if pts is None:
            return ""

        h, w = image_bgr.shape[:2]
        x1 = max(0, int(np.floor(np.min(pts[:, 0]) * w)) - 6)
        y1 = max(0, int(np.floor(np.min(pts[:, 1]) * h)) - 6)
        x2 = min(w - 1, int(np.ceil(np.max(pts[:, 0]) * w)) + 6)
        y2 = min(h - 1, int(np.ceil(np.max(pts[:, 1]) * h)) + 6)
        if x2 <= x1 or y2 <= y1:
            return ""

        crop = image_bgr[y1 : y2 + 1, x1 : x2 + 1]
        if crop.size == 0:
            return ""

        # 对小数值格做更激进的局部放大，优先救回 6.4 / 6.6 / 6.7 这类短文本。
        crop = cv2.copyMakeBorder(
            crop,
            8,
            8,
            8,
            8,
            borderType=cv2.BORDER_CONSTANT,
            value=(255, 255, 255),
        )
        crop = cv2.resize(crop, None, fx=2.4, fy=2.4, interpolation=cv2.INTER_CUBIC)

        try:
            items = self.ocr_service.recognize_text(crop, preprocess=True)
        except Exception:
            return ""
        if not items:
            return ""

        tokens = [str(item.text or "").strip() for item in items if str(item.text or "").strip()]
        if not tokens:
            return ""
        return "".join(tokens)

    @staticmethod
    def _repair_numeric_candidate(candidate: str, examples: Sequence[float]) -> Optional[int | float]:
        text = str(candidate or "").strip()
        if not text:
            return None
        normalized = re.sub(r"[^0-9.]", "", text)
        if not normalized:
            return None
        normalized = re.sub(r"\.{2,}", ".", normalized)
        if normalized.count(".") > 1:
            first = normalized.find(".")
            normalized = normalized[: first + 1] + normalized[first + 1 :].replace(".", "")

        parsed = FileParserService._parse_score(normalized)
        if parsed is not None:
            return parsed

        digits = re.sub(r"\D+", "", normalized)
        if not digits:
            return None

        if examples:
            decimals = [value for value in examples if abs(value - int(value)) > 1e-6]
            if decimals:
                median_example = float(np.median(np.array(decimals, dtype=np.float32)))
                if 0 < median_example < 10 and len(digits) >= 2:
                    repaired = f"{digits[0]}.{digits[1:]}"
                    parsed = FileParserService._parse_score(repaired)
                    if parsed is not None:
                        return parsed

        return None

    def _refine_col_bounds(
        self,
        headers: Sequence[str],
        col_bounds: List[float],
        header_boxes: Dict[str, Dict[str, Any]],
        rows_out: List[Dict[str, Any]],
        coord_key: str,
    ) -> List[float]:
        refined = list(col_bounds)
        for idx, header in enumerate(headers):
            boxes: List[Dict[str, Any]] = []
            if header in header_boxes:
                boxes.append(header_boxes[header])
            for row in rows_out:
                box = row.get("boxes", {}).get(header)
                if box:
                    boxes.append(box)
            if not boxes:
                continue

            x_ranges = []
            for box in boxes:
                pts = self._points_array(box, coord_key)
                if pts is None:
                    continue
                xs = pts[:, 0]
                x_ranges.append((float(xs.min()), float(xs.max())))
            if not x_ranges:
                continue

            min_x = min(r[0] for r in x_ranges)
            max_x = max(r[1] for r in x_ranges)
            width = max_x - min_x
            pad_left = width * (0.62 if idx == 0 else 0.34)
            pad_right = width * (0.42 if self._is_numeric_column(header) else 0.48)

            candidate_left = min_x - pad_left
            candidate_right = max_x + pad_right

            # 保持列边界单调，避免相邻列反向交叉
            refined[idx] = min(refined[idx], candidate_left)
            refined[idx + 1] = max(refined[idx + 1], candidate_right)

        # 再做一次单调性修正
        for i in range(1, len(refined)):
            if refined[i] <= refined[i - 1]:
                refined[i] = refined[i - 1] + 0.01
        return refined

    def _refine_row_bounds(
        self,
        row_bounds: List[float],
        header_boxes: Dict[str, Dict[str, Any]],
        rows_out: List[Dict[str, Any]],
        coord_key: str,
    ) -> List[float]:
        refined = list(row_bounds)
        row_boxes: List[List[Dict[str, Any]]] = []
        header_row_boxes = list(header_boxes.values())
        row_boxes.append(header_row_boxes)
        for row in rows_out:
            row_boxes.append(list(row.get("boxes", {}).values()))

        for idx, boxes in enumerate(row_boxes):
            if not boxes:
                continue
            y_ranges = []
            for box in boxes:
                pts = self._points_array(box, coord_key)
                if pts is None:
                    continue
                ys = pts[:, 1]
                y_ranges.append((float(ys.min()), float(ys.max())))
            if not y_ranges:
                continue

            min_y = min(r[0] for r in y_ranges)
            max_y = max(r[1] for r in y_ranges)
            height = max_y - min_y
            pad_top = height * (0.46 if idx == 0 else 0.36)
            pad_bottom = height * 0.42

            refined[idx] = min(refined[idx], min_y - pad_top)
            refined[idx + 1] = max(refined[idx + 1], max_y + pad_bottom)

        for i in range(1, len(refined)):
            if refined[i] <= refined[i - 1]:
                refined[i] = refined[i - 1] + 0.01
        return refined

    @staticmethod
    def _is_numeric_column(header: str) -> bool:
        return any(token in header for token in ("成绩", "分数", "学分", "班级", "学号", "平时", "期中", "期末"))

    @staticmethod
    def _estimate_table_axes(
        boxes: Sequence[Dict[str, Any]],
        coord_key: str = "points",
    ) -> Tuple[Optional[np.ndarray], Optional[np.ndarray], Optional[np.ndarray]]:
        x_vectors: List[np.ndarray] = []
        y_vectors: List[np.ndarray] = []
        for box in boxes:
            pts = FileParserService._points_array(box, coord_key)
            if pts is None or len(pts) != 4:
                continue
            top = pts[1] - pts[0]
            left = pts[3] - pts[0]
            top_norm = np.linalg.norm(top)
            left_norm = np.linalg.norm(left)
            if top_norm > 1e-6:
                x_vectors.append(top / top_norm)
            if left_norm > 1e-6:
                y_vectors.append(left / left_norm)
        if not x_vectors or not y_vectors:
            return None, None, None
        x_axis = np.mean(np.stack(x_vectors), axis=0)
        y_axis = np.mean(np.stack(y_vectors), axis=0)
        x_axis /= max(np.linalg.norm(x_axis), 1e-6)
        y_axis /= max(np.linalg.norm(y_axis), 1e-6)
        origins: List[np.ndarray] = []
        for box in boxes:
            pts = FileParserService._points_array(box, coord_key)
            if pts is None or len(pts) != 4:
                continue
            center = np.mean(pts, axis=0)
            x_pos = float(np.dot(center, x_axis))
            y_pos = float(np.dot(center, y_axis))
            origins.append(center - x_axis * x_pos - y_axis * y_pos)
        if not origins:
            return None, None, None
        origin = np.median(np.stack(origins), axis=0)
        return origin.astype(np.float32), x_axis, y_axis

    @staticmethod
    def _points_array(box: Dict[str, Any], coord_key: str = "points") -> Optional[np.ndarray]:
        points = box.get(coord_key)
        if not points:
            return None
        return np.array([[float(p["x"]), float(p["y"])] for p in points], dtype=np.float32)

    @staticmethod
    def _polygon_center(box: Dict[str, Any], coord_key: str = "points") -> np.ndarray:
        pts = FileParserService._points_array(box, coord_key)
        return np.mean(pts, axis=0)

    @staticmethod
    def _safe_median(values: Sequence[float]) -> Optional[float]:
        if not values:
            return None
        return float(np.median(np.array(values, dtype=np.float32)))

    @staticmethod
    def _positions_to_bounds(positions: Sequence[float]) -> List[float]:
        if len(positions) == 1:
            half = 0.05
            return [positions[0] - half, positions[0] + half]
        bounds: List[float] = []
        first_gap = (positions[1] - positions[0]) / 2.0
        bounds.append(positions[0] - first_gap)
        for left, right in zip(positions, positions[1:]):
            bounds.append((left + right) / 2.0)
        last_gap = (positions[-1] - positions[-2]) / 2.0
        bounds.append(positions[-1] + last_gap)
        return bounds

    @staticmethod
    def _bounds_to_polygon(
        origin: np.ndarray,
        x_axis: np.ndarray,
        y_axis: np.ndarray,
        left: float,
        right: float,
        top: float,
        bottom: float,
        transform_context: Optional[Dict[str, Any]],
        coord_key: str = "points",
    ) -> Dict[str, Any]:
        points = np.array([
            origin + x_axis * left + y_axis * top,
            origin + x_axis * right + y_axis * top,
            origin + x_axis * right + y_axis * bottom,
            origin + x_axis * left + y_axis * bottom,
        ], dtype=np.float32)
        if transform_context is None or coord_key == "points":
            normalized_points = [
                {"x": round(float(p[0]), 6), "y": round(float(p[1]), 6)}
                for p in points
            ]
            result: Dict[str, Any] = {"points": normalized_points}
        else:
            result = {
                "processed_points": [
                    {"x": round(float(p[0]), 6), "y": round(float(p[1]), 6)}
                    for p in points
                ]
            }
            points = FileParserService._project_points_to_original(points, transform_context)
            result["points"] = [
                {"x": round(float(p[0]), 6), "y": round(float(p[1]), 6)}
                for p in points
            ]
        if transform_context is not None and coord_key == "points":
            original_size = transform_context["original_size"]
            original_width = max(1, int(original_size["width"]))
            original_height = max(1, int(original_size["height"]))
            absolute_points = points.copy()
            absolute_points[:, 0] *= original_width
            absolute_points[:, 1] *= original_height
            processed_points = FileParserService._project_points_to_processed(
                absolute_points,
                transform_context,
            )
            result["processed_points"] = processed_points
        return result

    @staticmethod
    def _project_points_to_original(points: np.ndarray, transform_context: Dict[str, Any]) -> np.ndarray:
        inverse_matrix = np.array(transform_context["inverse_matrix"], dtype=np.float32)
        processed_size = transform_context["processed_size"]
        original_size = transform_context["original_size"]
        processed_width = max(1, int(processed_size["width"]))
        processed_height = max(1, int(processed_size["height"]))
        original_width = max(1, int(original_size["width"]))
        original_height = max(1, int(original_size["height"]))
        absolute_points = points.copy()
        absolute_points[:, 0] *= processed_width
        absolute_points[:, 1] *= processed_height
        projected = cv2.perspectiveTransform(absolute_points.reshape(1, -1, 2), inverse_matrix)[0]
        projected[:, 0] /= original_width
        projected[:, 1] /= original_height
        return projected

    @staticmethod
    def _project_points_to_processed(points: np.ndarray, transform_context: Dict[str, Any]) -> List[Dict[str, float]]:
        forward_matrix = np.array(transform_context["forward_matrix"], dtype=np.float32)
        processed_size = transform_context["processed_size"]
        width = max(1, int(processed_size["width"]))
        height = max(1, int(processed_size["height"]))
        projected = cv2.perspectiveTransform(points.reshape(1, -1, 2), forward_matrix)[0]
        return [
            {"x": round(float(p[0] / width), 6), "y": round(float(p[1] / height), 6)}
            for p in projected
        ]

    @staticmethod
    def _normalize_polygon(
        box: Optional[List[List[float]]],
        image_size: Optional[Dict[str, int]],
        transform_context: Optional[Dict[str, Any]],
    ) -> Optional[Dict[str, Any]]:
        if not box or not image_size:
            return None
        width = max(1, int(image_size.get("width", 0)))
        height = max(1, int(image_size.get("height", 0)))
        points = np.array(box, dtype=np.float32)
        if transform_context is not None:
            inverse_matrix = np.array(transform_context["inverse_matrix"], dtype=np.float32)
            points = cv2.perspectiveTransform(points.reshape(1, -1, 2), inverse_matrix)[0]
        normalized_points = [
            {
                "x": round(float(p[0] / width), 6),
                "y": round(float(p[1] / height), 6),
            }
            for p in points
        ]
        result: Dict[str, Any] = {"points": normalized_points}
        if transform_context is not None:
            processed_points = FileParserService._project_points_to_processed(points, transform_context)
            result["processed_points"] = processed_points
        return result

    def _extract_from_text_lines_fallback(
        self,
        lines: List[str],
        *,
        meta: Optional[Dict[str, Any]] = None,
        image_size: Optional[Dict[str, int]] = None,
        ocr_items: Optional[Sequence[OCRItem]] = None,
    ) -> Dict[str, Any]:
        """
        当无法识别表格结构时，返回一个简单的单列数据
        """
        rows = []
        for line in lines:
            if not line.strip():
                continue
            row = {
                "values": {"内容": line.strip()},
                "confidences": {"内容": 0.8} # 默认低置信度提示用户检查
            }
            if ocr_items:
                box = OCRService.best_match_box(line.strip(), ocr_items)
                normalized_box = self._normalize_polygon(box, image_size, transform_context=None)
                if normalized_box is not None:
                    row["boxes"] = {"内容": normalized_box}
            rows.append(row)
        result = {
            "headers": ["内容"],
            "rows": rows
        }
        if meta:
            result["meta"] = meta
            if meta.get("标题"):
                result["title"] = meta["标题"]
        return result

    @staticmethod
    def _parse_score(v: Any) -> Optional[int | float]:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return int(v) if isinstance(v, int) or abs(float(v) - int(float(v))) < 1e-9 else float(v)

        s = str(v).strip()
        # 仅匹配纯数字，或者带小数点的数字
        if not re.match(r"^-?\d+(\.\d+)?$", s):
            return None
            
        try:
            num = float(s)
            return int(num) if abs(num - int(num)) < 1e-9 else round(num, 2)
        except Exception:
            return None
