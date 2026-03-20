import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from backend.services.excel_service import ExcelService


class ExcelServiceTests(unittest.TestCase):
    def test_export_transcript_creates_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            service = ExcelService(export_dir=Path(tmpdir))
            filename = service.export_transcript(
                {
                    "title": "学生成绩单",
                    "headers": ["姓名", "课程", "成绩"],
                    "rows": [{"values": {"姓名": "张三", "课程": "数学", "成绩": "95"}, "confidences": {}}],
                }
            )

            path = Path(tmpdir) / filename
            self.assertTrue(path.exists())

            wb = load_workbook(path)
            ws = wb.active
            self.assertEqual(ws["A1"].value, "学生成绩单")
            self.assertEqual(ws["A2"].value, "姓名")
            self.assertEqual(ws["C3"].value, 95)


if __name__ == "__main__":
    unittest.main()
